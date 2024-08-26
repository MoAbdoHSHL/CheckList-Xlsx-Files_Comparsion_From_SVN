import os
import subprocess
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import xlwings as xw

def create_update_file_list(excel_path):
    if os.path.exists(excel_path):
        workbook = load_workbook(excel_path)
    else:
        workbook = Workbook()
        workbook.remove(workbook.active)  # Remove default sheet
    
    # Create 'FileList' sheet or clear existing content
    if 'FileList' in workbook.sheetnames:
        file_list_sheet = workbook['FileList']
        file_list_sheet.delete_rows(2, file_list_sheet.max_row)  # Clear existing data
    else:
        file_list_sheet = workbook.create_sheet('FileList', 0)
    
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='4682B4', end_color='4682B4', fill_type='solid')  
    
    headers = ['Files from Package_TAGS_Main', 'ReviewCycleDate', 'Path', 'VersionReview', 
                'ReviewCycleDate', 'Path', 'VersionReview', 'Compare']
    
    column_widths = {'A': 15, 'B': 15, 'C': 40, 'D': 15, 'E': 20, 'F': 40, 'G': 15}

    for col, header in enumerate(headers, start=1):
        cell = file_list_sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        column_letter = cell.column_letter
        if column_letter in column_widths:
            file_list_sheet.column_dimensions[column_letter].width = column_widths[column_letter]
    
    file_list_sheet.row_dimensions[1].height = 30

    # Apply cell formatting for data rows
    for row in file_list_sheet.iter_rows(min_row=2, max_col=8):
        for cell in row:
            if cell.column in [3, 6]:  # Columns C and F
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # Formatting dates in yyyy-mm-dd format
    for row in file_list_sheet.iter_rows(min_row=2, max_col=8):
        for cell in row:
            if cell.column in [2, 5]:  # Columns B and E
                if cell.value:
                    if isinstance(cell.value, str) and '-' in cell.value:
                        try:
                            cell.value = cell.value[:10]  # Ensure yyyy-mm-dd format
                        except ValueError:
                            pass
                    elif isinstance(cell.value, (int, float)):
                        try:
                            cell.value = cell.value.strftime('%Y-%m-%d')
                        except AttributeError:
                            pass

    workbook.save(excel_path)


def add_vba_code_and_button(excel_path):
    # Open the workbook with xlwings
    wb = xw.Book(excel_path)
    
    # Select the sheet where the button will be added
    ws = wb.sheets['FileList']
    
    # Define the VBA macro code as a string
    vba_code = """
Sub CompareVersions()
    Dim svnUrl1 As String
    Dim svnUrl2 As String
    Dim ws As Worksheet
    Dim fileInfo1 As Variant
    Dim fileInfo2 As Variant
    Dim fileName As Variant
    Dim row As Integer
    Dim fileList As Collection

    On Error GoTo ErrorHandler

    ' Initialize the worksheet
    Set ws = ThisWorkbook.Sheets("FileList")
    If ws Is Nothing Then
        MsgBox "Sheet 'FileList' not found!", vbCritical
        Exit Sub
    End If

    ' Prompt the user to enter the SVN URLs
    svnUrl1 = InputBox("Enter the first SVN URL:", "SVN URL 1")
    If svnUrl1 = "" Then Exit Sub

    svnUrl2 = InputBox("Enter the second SVN URL:", "SVN URL 2")
    If svnUrl2 = "" Then Exit Sub

    ' Get file information from both SVN URLs
    Set fileList = GetFilesInSVN(svnUrl1, "c,h")
    row = 2
    For Each fileName In fileList
        Set fileInfo1 = GetFileInfo(svnUrl1, CStr(fileName))
        Set fileInfo2 = GetFileInfo(svnUrl2, CStr(fileName))
        
        ' Display file information
        ws.Cells(row, 1).Value = CStr(fileName)
        
        ' Date and revision
        ws.Cells(row, 2).Value = IIf(Not fileInfo1 Is Nothing, fileInfo1("date"), "Not found")
        ws.Cells(row, 4).Value = IIf(Not fileInfo1 Is Nothing, fileInfo1("revision"), "Not found")
        ws.Cells(row, 5).Value = IIf(Not fileInfo2 Is Nothing, fileInfo2("date"), "Not found")
        ws.Cells(row, 7).Value = IIf(Not fileInfo2 Is Nothing, fileInfo2("revision"), "Not found")
        
        ' Hyperlinks
        If Not fileInfo1 Is Nothing Then
            ws.Hyperlinks.Add Anchor:=ws.Cells(row, 3), Address:=svnUrl1 & "/" & CStr(fileName), TextToDisplay:=svnUrl1 & "/" & CStr(fileName)
        Else
            ws.Cells(row, 3).Value = "Not found"
        End If
        If Not fileInfo2 Is Nothing Then
            ws.Hyperlinks.Add Anchor:=ws.Cells(row, 6), Address:=svnUrl2 & "/" & CStr(fileName), TextToDisplay:=svnUrl2 & "/" & CStr(fileName)
        Else
            ws.Cells(row, 6).Value = "Not found"
        End If
        
        ' Compare the file versions
        If Not fileInfo1 Is Nothing And Not fileInfo2 Is Nothing Then
            If fileInfo1("revision") = fileInfo2("revision") Then
                ws.Cells(row, 8).Value = "Match"
            Else
                ws.Cells(row, 8).Value = "Mismatch"
            End If
        Else
            ws.Cells(row, 8).Value = "Not found"
        End If
        
        row = row + 1
    Next fileName

    ' Set column widths
    ws.Columns("A:A").AutoFit
    ws.Columns("B:B").AutoFit
    ws.Columns("D:D").AutoFit
    ws.Columns("E:E").AutoFit
    ws.Columns("G:G").AutoFit
    ws.Columns("H:H").AutoFit
    ws.Columns("I:I").AutoFit

    ' Set fixed width for columns C and F
    ws.Columns("C:C").ColumnWidth = 40 
    ws.Columns("F:F").ColumnWidth = 40 
    Exit Sub

    ' Set alignment for header row
    With ws.Rows(1)
        .HorizontalAlignment = xlCenter
        .VerticalAlignment = xlCenter
        .Font.Bold = True
        .RowHeight = 30
    End With
    
    ' Set alignment for all cells
    With ws.Range("A2:H" & row - 1)
        .HorizontalAlignment = xlCenter
        .VerticalAlignment = xlCenter
    End With
    
    ' Special alignment for columns C and F
    ws.Columns("C:C").HorizontalAlignment = xlLeft
    ws.Columns("F:F").HorizontalAlignment = xlLeft

    Exit Sub
    
ErrorHandler:
    MsgBox "An error occurred: " & Err.Description, vbCritical
End Sub

Function GetFilesInSVN(svnUrl As String, fileExts As String) As Collection
    Dim shell As Object
    Dim exec As Object
    Dim output As String
    Dim lines As Variant
    Dim fileList As New Collection
    Dim line As Variant
    Dim exts As Variant
    Dim ext As Variant

    ' Initialize shell object
    Set shell = CreateObject("WScript.Shell")

    ' Run SVN list command
    Set exec = shell.Exec("svn list -R " & svnUrl)
    Do While Not exec.StdOut.AtEndOfStream
        output = output & exec.StdOut.ReadLine & vbCrLf
    Loop

    ' Split output into lines
    lines = Split(output, vbCrLf)
    exts = Split(fileExts, ",")

    ' Filter files by extension
    For Each line In lines
        For Each ext In exts
            If LCase(Right(line, Len(ext) + 1)) = "." & LCase(ext) Then
                fileList.Add line
                Exit For
            End If
        Next ext
    Next line

    Set GetFilesInSVN = fileList
End Function

Function GetFileInfo(svnUrl As String, fileName As String) As Variant
    Dim shell As Object
    Dim exec As Object
    Dim output As String
    Dim lines As Variant
    Dim line As Variant
    Dim info As Object
    Dim datePart As String
    Dim revisionPart As String

    ' Initialize shell object
    Set shell = CreateObject("WScript.Shell")

    ' Run SVN info command
    Set exec = shell.Exec("svn info " & svnUrl & "/" & fileName)
    Do While Not exec.StdOut.AtEndOfStream
        output = output & exec.StdOut.ReadLine & vbCrLf
    Loop

    ' Split output into lines
    lines = Split(output, vbCrLf)

    ' Parse SVN info
    Set info = CreateObject("Scripting.Dictionary")
    For Each line In lines
        If InStr(line, "Last Changed Date:") > 0 Then
            datePart = Trim(Split(line, "Last Changed Date:")(1))
            info.Add "date", datePart
        ElseIf InStr(line, "Last Changed Rev:") > 0 Then
            revisionPart = Trim(Split(line, "Last Changed Rev:")(1))
            info.Add "revision", revisionPart
        End If
    Next line

    If info.Exists("date") And info.Exists("revision") Then
        Set GetFileInfo = info
    Else
        Set GetFileInfo = Nothing
    End If
End Function

    """
    
    # Add VBA code to the workbook
    wb.api.VBProject.VBComponents.Add(1).CodeModule.AddFromString(vba_code)
    button = ws.api.Buttons().Add(400, 120, 100, 50)  
    button.Text = "Compare"
    button.OnAction = "CompareVersions"
    
    wb.save(excel_path)

def main():
    excel_path = r'D:local file path.xlsx'
    
    # Create or update the 'FileList' sheet
    create_update_file_list(excel_path)
    
    # Add VBA code and button to the sheet
    add_vba_code_and_button(excel_path)

if __name__ == "__main__":
    main()
